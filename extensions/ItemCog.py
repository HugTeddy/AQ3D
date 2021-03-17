import discord
from discord.ext import tasks, commands
import openpyxl
import re
from disputils import BotEmbedPaginator

itemlist=[]
regexmatch=["[[", "]]"]
wb = openpyxl.load_workbook('../data/AQ3DItemDatabase.xlsx')
ws = wb['All']
for col in ws['B']:
    itemlist.append(col.value.lower())

def getItem(row:int):
    global ws
    data=[]
    for item in ws[row+1]:
        data.append(item.value)
    return data

async def paginate(ctx, embeds):
    paginator = BotEmbedPaginator(ctx, embeds)
    await paginator.run()

def createItemEmbed(data):
    attributes = ""
    index = [7,8,9,10,11,12,13]
    index_name = {"7":"Armor","8":"Evasion","9":"Haste","10":"Health","11":"Attack","12":"Critical","13":"Stack Size"}
    colors = {"Common":discord.Color.light_grey(), "Uncommon": discord.Color.blue(), "Epic": discord.Color.orange(), "Awesome": discord.Color.red(), "Junk": discord.Color.blurple(), "Legendary": discord.Color.from_rgb(255,255,0)}
    if data[6] in colors.keys():
        color = colors[data[6]]
    else:
        color = discord.Color.default()
    for i in index:
        if data[i]:
            attributes += f' {index_name[str(i)]} +{format(int(data[i]), ",d")}'
    embed = discord.Embed(
        title=data[1],
        description=f'{data[2]}\nItem ID: {int(data[0])}',
        color=color
    )
    embed.add_field(name='Type:', value=f'{data[5]}', inline=True)
    embed.add_field(name='Level:', value=f'{format(int(data[3]), ",d")}', inline=True)
    embed.add_field(name='Rarity:', value=f'{data[6]}', inline=True)
    if data[4]:
        embed.add_field(name='Value:',value=f'{format(int(data[4]), ",d")}', inline=True)
    else:
        embed.add_field(name='Value:',value=f'0', inline=True)
    embed.add_field(name='Attributes:', value=f'{attributes}', inline=True)
    if data[14] == "TRUE":
        embed.add_field(name='Cosmetic:', value=f'✅', inline=True)
    else:
        embed.add_field(name='Cosmetic:', value=f'❌', inline=True)
    return embed

class ItemCog(commands.Cog):
    def __init__(self, client):
        self.client = client

    @commands.Cog.listener()
    async def on_message(self, message):
        global itemlist
        global regexmatch
        embeds = []
        msg = message.content
        ctx = await self.client.get_context(message)
        items = []
        if all(x in msg for x in regexmatch):
            items = re.findall('(?<=\[\[)(.*?)(?=\]\])', msg)
            for iName in items:
                iName = iName.lower()
                if iName in itemlist:
                    row = itemlist.index(iName)
                    data = getItem(row)
                    embed = createItemEmbed(data)
                    embeds.append(embed)
            if len(embeds) == 0:
                return
            if len(embeds) == 1:
                await message.channel.send(embed=embeds[0])
            else:
                await paginate(ctx, embeds)

def setup(client):
    client.add_cog(ItemCog(client))
